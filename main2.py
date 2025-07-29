from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook
import os
import uuid
import re
import logging
from fastapi.middleware.cors import CORSMiddleware

# Setup logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class TemplateData(BaseModel):
    replacements: dict[str, str]

def replace_placeholders_in_sheet(sheet, replacements):
    pattern = re.compile(r'\((\d+)\)')  # matches (1), (2), ...
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                matches = pattern.findall(cell.value)
                if matches:
                    original = cell.value
                    for match in matches:
                        if match in replacements:
                            original = original.replace(f"({match})", str(replacements[match]))
                        else:
                            logger.warning(f"Missing replacement for ({match})")
                    cell.value = original

@app.post("/generate-excel-2/")
async def generate_excel_2(data: TemplateData):
    try:
        template_path = "Copy of Custom_Doc-DOCUMENTS_5855_TEMPLATE.xlsx"
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="Template not found")

        wb = load_workbook(template_path)

        for sheet in wb.worksheets:
            replace_placeholders_in_sheet(sheet, data.replacements)

        os.makedirs("temp", exist_ok=True)
        output_path = os.path.join("temp", f"output_{uuid.uuid4()}.xlsx")
        wb.save(output_path)

        return FileResponse(
            path=output_path,
            filename=os.path.basename(output_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error("Error generating Excel", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal error: {str(e)}")
