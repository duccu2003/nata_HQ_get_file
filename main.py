from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook
import os
import uuid
from typing import Dict
import re
import logging
from flatten_dict import flatten
from fastapi.middleware.cors import CORSMiddleware
from demo import fill_invoice_template
from demo import get_invoice_data

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
@app.get("/")
def read_root():
    return {"message": "Welcome to FastAPI!"}

from fastapi import FastAPI
from fastapi.responses import FileResponse

app = FastAPI()

@app.get("/get_invoice_data/")
def get_data(contract_number: str):
    result = get_invoice_data(contract_number)
    if not result:
        raise HTTPException(status_code=404, detail="Contract not found")
    return result

@app.get("/export-invoice/")
def export_invoice(contract_number: str):
    template = "Copy of Custom_Doc-DOCUMENTS_5855_TEMPLATE.xlsx"
    output = f"invoice_{contract_number.replace('/', '_')}.xlsx"
    
    try:
        path = fill_invoice_template(contract_number, template, output)
        return FileResponse(path, filename=output)
    except Exception as e:
        return {"error": str(e)}


class TemplateData(BaseModel):
    replacements: Dict[str, str]

def replace_placeholders_in_sheet(sheet, replacements):
    
    placeholder_pattern = re.compile(r'\(\d+\)')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                matches = placeholder_pattern.findall(cell.value)
                if matches:
                    new_value = cell.value
                    for match in matches:
                        placeholder_key = match[1:-1]
                        if placeholder_key in replacements:
                            new_value = new_value.replace(match, replacements[placeholder_key])
                    cell.value = new_value

@app.post("/generate-excel/")
async def generate_excel(data: TemplateData):
    try:
        template_path = "Copy of Custom_Doc-DOCUMENTS_5855_TEMPLATE.xlsx"
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="Template file not found")

        wb = load_workbook(template_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            replace_placeholders_in_sheet(sheet, data.replacements)
        
        output_filename = f"output_{uuid.uuid4()}.xlsx"
        output_path = os.path.join("temp", output_filename)
        os.makedirs("temp", exist_ok=True)
        wb.save(output_path)
        
        # return FileResponse(
        #     path=output_path,
        #     filename=output_filename,
        #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # )
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel file: {str(e)}")

class TemplateData(BaseModel):
    replacements: Dict[str, str] | None = None

def replace_placeholders_in_sheet(sheet, replacements):
    placeholder_pattern = re.compile(r'\(\d+\)')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                matches = placeholder_pattern.findall(cell.value)
                if matches:
                    new_value = cell.value
                    for match in matches:
                        placeholder_key = match[1:-1]
                        if placeholder_key in replacements:
                            new_value = new_value.replace(match, str(replacements[placeholder_key]))
                        else:
                            logger.warning(f"No replacement found for placeholder: {match}")
                    cell.value = new_value

@app.post("/generate-excel-2/")
async def generate_excel_2(data: dict):
    try:
        template_path = "Copy of Custom_Doc-DOCUMENTS_5855_TEMPLATE.xlsx"
        logger.debug(f"Template path: {os.path.abspath(template_path)}")

        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="Template file not found")

        wb = load_workbook(template_path)
        flat_replacements = {}

        # Maps ID to Name for seller and buyer
        SELLER_NAME_MAP = {
            1: "SUNWOOD VIETNAM",
            2: "ACME EXPORT",
            # add more if needed
        }

        BUYER_NAME_MAP = {
            2: "KOREA BIOMASS CO., LTD.",
            3: "ABC ENERGY",
            # add more if needed
        }

        if "replacements" in data and isinstance(data["replacements"], dict):
            flat_replacements = {k: str(v or "") for k, v in data["replacements"].items()}
        else:
            flat_dict = flatten(data, reducer='dot')
            logger.debug(f"Flattened input: {flat_dict}")

            # Handle SELLER_ID and BUYER_ID by name
            seller_id = flat_dict.get('T_PURCHASE_CONTRACT.SELLER_ID')
            if seller_id in SELLER_NAME_MAP:
                flat_replacements['1'] = SELLER_NAME_MAP[seller_id]

            buyer_id = flat_dict.get('T_PURCHASE_CONTRACT.BUYER_ID')
            if buyer_id in BUYER_NAME_MAP:
                flat_replacements['4'] = BUYER_NAME_MAP[buyer_id]

            placeholder_map = {
                'T_PURCHASE_CONTRACT.CODE': '8',
                'T_PURCHASE_CONTRACT.CONTRACT_DATE': '3',
                'T_PURCHASE_CONTRACT.PORT_OF_LOADING': '11',
                'T_PURCHASE_CONTRACT.PORT_OF_DISCHARGE': '12',
                'T_PURCHASE_CONTRACT.VESSEL_NAME': '13',
                'T_PURCHASE_CONTRACT.ETD_DATE': '14',
                'T_PURCHASE_CONTRACT.COMMODITY': '16',
                'T_PURCHASE_CONTRACT.ORIGIN': '17',
                'T_PURCHASE_CONTRACT.HS_CODE': '18',
                'T_PURCHASE_CONTRACT.SHIPMENT_TERM': '19',
                'T_PURCHASE_CONTRACT.UNIT_PRICE': '20',
                'T_PURCHASE_CONTRACT.NET_WEIGHT': '21',
                'T_PURCHASE_CONTRACT.GROSS_WEIGHT': '22',
                'T_PURCHASE_CONTRACT.NUMBER_OF_CONTAINERS': '23',
                'T_PURCHASE_CONTRACT.LC_NO': '24',
                'T_PURCHASE_CONTRACT.LC_DATE': '25',
                'T_PURCHASE_CONTRACT_WEIGHT_TICKET_DETAIL.NET_WEIGHT': '21',
                'T_PURCHASE_CONTRACT_WEIGHT_TICKET_DETAIL.GROSS_WEIGHT': '22',
                'T_PURCHASE_CONTRACT_WEIGHT_TICKET_DETAIL.CONTAINER_NUMBER': '30',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.SHIP_NAME': '13',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.ETD_DATE': '14',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.EXPORT_PORT': '11',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.IMPORT_PORT': '12',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.CONTAINER_QUANTITY': '23',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.BOOKING_NUMBER': '28',
                'T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE.ETA_DATE': '29',
                'T_PURCHASE_CONTRACT_GOOD.GOOD_TYPE': '16',
                'T_PURCHASE_CONTRACT_GOOD.QUANTITY': '21',
                'T_PURCHASE_CONTRACT_GOOD.UNIT': '27',
                'T_PURCHASE_CONTRACT_GOOD.HS_CODE': '18',
                # Alias or reused placeholders
                'T_PURCHASE_CONTRACT.CODE': '2',
                'T_PURCHASE_CONTRACT.SHIPMENT_TERM': '10',
                'T_PURCHASE_CONTRACT.LC_NO': '9',
            }

            for flat_key, flat_value in flat_dict.items():
                if flat_key in placeholder_map:
                    placeholder = placeholder_map[flat_key]
                    flat_replacements[placeholder] = str(flat_value or "")

            # Add default empty values
            flat_replacements.update({
                '5': '',  # Consignee Rep
                '6': '',  # Applicant Rep
                '7': '',  # Consignee Phone
                '15': ''  # Remarks
            })

            logger.debug(f"Final replacements: {flat_replacements}")

        # Calculate GRAND TOTAL = Unit Price * Quantity
        if '21' in flat_replacements and '20' in flat_replacements:
            try:
                quantity = float(flat_replacements.get('21', 0))
                unit_price = float(flat_replacements.get('20', 0))
                grand_total = quantity * unit_price
                flat_replacements['26'] = f"USD {grand_total:,.2f}"
                logger.debug(f"Calculated GRAND TOTAL (26): {flat_replacements['26']}")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid number for quantity or unit price")

        # Apply replacements
        for sheet_name in wb.sheetnames:
            logger.debug(f"Processing sheet: {sheet_name}")
            sheet = wb[sheet_name]
            replace_placeholders_in_sheet(sheet, flat_replacements)

        os.makedirs("temp", exist_ok=True)
        if not os.access("temp", os.W_OK):
            raise HTTPException(status_code=500, detail="No write permission for temp folder")

        output_filename = f"output_{uuid.uuid4()}.xlsx"
        output_path = os.path.join("temp", output_filename)
        wb.save(output_path)

        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")