# from openpyxl import load_workbook
# import re
# from sqlalchemy import create_engine, text
# from datetime import datetime
# from num2words import num2words
# import json
# import mysql.connector
# import mysql.connector
# from datetime import datetime
# from num2words import num2words

# def format_date(dt, fmt='%d/%m/%Y'):
#     return dt.strftime(fmt) if dt else None

# def format_amount_text(amount: float) -> str:
#     if amount is None:
#         return ""
#     text_en = num2words(amount, to='currency', lang='en')
#     return f"US DOLLARS {text_en.upper()} ONLY"

# def get_invoice_data(contract_number: str):
#     # Kết nối MySQL
#     conn = mysql.connector.connect(
#         host="localhost",       # hoặc IP nếu khác
#         port=4306,
#         user="training",
#         password="123%4056789",   # không cần mã hóa như SQLAlchemy
#         database="smartwood"
#     )
#     cursor = conn.cursor(dictionary=True)

#     query = """
#         SELECT 
#             pc.CONTRACT_NUMBER,
#             pc.CONTRACT_DATE,
#             CONCAT('INPUT ', pc.CONTRACT_NUMBER) AS CONTRACT_TITLE,

#             sc.LC_CONTRACT_NUMBER,
#             sc.LC_DATE,
#             sc.PAYMENT_TERM,
#             sc.GOOD_DESCRIPTION AS SHIPMENT_TERM,
#             sc.TOTAL_WEIGHT,

#             cust.NAME AS BUYER_NAME,
#             cust.ADDRESS AS BUYER_ADDRESS,
#             cust.REPRESENTED AS LEGAL_REPRESENTATIVE,
#             cust.PHONE,

#             ss.EXPORT_PORT,
#             ss.IMPORT_PORT,
#             ss.SHIP_NAME,
#             ss.ETD_DATE,
#             ss.AVAILABLE_CONTAINER_QUANTITY,

#             sg.UNIT_PRICE,
#             g.DESCRIPTION AS COMMODITY,
#             g.HS_CODE,
#             g.ORIGIN_COUNTRY

#         FROM T_PURCHASE_CONTRACT pc
#         LEFT JOIN T_SALE_CONTRACT sc ON sc.CONTRACT_ID_INT = pc.ID
#         LEFT JOIN M_CUSTOMER cust ON sc.CUSTOMER_ID_INT = cust.ID
#         LEFT JOIN T_SALE_CONTRACT_GOOD sg ON sg.SALE_CONTRACT_ID_INT = sc.ID
#         LEFT JOIN M_GOOD g ON sg.GOOD_ID_INT = g.ID
#         LEFT JOIN T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE ss ON ss.PURCHASE_CONTRACT_ID_INT = pc.ID
#         WHERE pc.CONTRACT_NUMBER = %s
#         LIMIT 1
#     """

#     cursor.execute(query, (contract_number,))
#     result = cursor.fetchone()
#     cursor.close()
#     conn.close()

#     if not result:
#         return None

#     net_weight = float(result["TOTAL_WEIGHT"] or 0)
#     gross_weight = round(net_weight + 0.4, 2)
#     unit_price = float(result["UNIT_PRICE"] or 0)
#     amount = round(unit_price * net_weight, 2)
#     containers = f"{int(result['AVAILABLE_CONTAINER_QUANTITY'])}x20'DC"

#     return {
#         "replacements": {
#             "1": result["CONTRACT_TITLE"],
#             "2": result["CONTRACT_NUMBER"],
#             "3": format_date(result["CONTRACT_DATE"]),
#             "4": result["BUYER_NAME"],
#             "5": result["BUYER_ADDRESS"],
#             "6": result["LEGAL_REPRESENTATIVE"],
#             "7": result["PHONE"],
#             "8": result["CONTRACT_NUMBER"],
#             "9": result["PAYMENT_TERM"],
#             "10": result["SHIPMENT_TERM"],
#             "11": result["EXPORT_PORT"],
#             "12": result["IMPORT_PORT"],
#             "13": result["SHIP_NAME"],
#             "14": format_date(result["ETD_DATE"]),
#             "15": format_amount_text(amount),
#             "16": result["COMMODITY"],
#             "17": result["ORIGIN_COUNTRY"],
#             "18": result["HS_CODE"],
#             "19": result["SHIPMENT_TERM"],
#             "20": unit_price,
#             "21": net_weight,
#             "22": gross_weight,
#             "23": containers,
#             "24": result["LC_CONTRACT_NUMBER"],
#             "25": format_date(result["LC_DATE"]),
#             "26": format_amount_text(amount)
#         }
#     }
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
import re
from datetime import datetime
from num2words import num2words

# Hàm format ngày
def format_date(dt, fmt='%d/%m/%Y'):
    return dt.strftime(fmt) if dt else None

# Hàm đổi số thành chữ
def format_amount_text(amount: float) -> str:
    if amount is None:
        return ""
    text_en = num2words(amount, to='currency', lang='en')
    return f"US DOLLARS {text_en.upper()} ONLY"

# Cấu hình engine SQLAlchemy
DATABASE_URL = "mysql+pymysql://training:123%4056789@localhost:4306/smartwood"
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def get_invoice_data(contract_number: str):
    session = SessionLocal()
    try:
        query = text("""
            SELECT 
                pc.CONTRACT_NUMBER,
                pc.CONTRACT_DATE,
                CONCAT('INPUT ', pc.CONTRACT_NUMBER) AS CONTRACT_TITLE,

                sc.LC_CONTRACT_NUMBER,
                sc.LC_DATE,
                sc.PAYMENT_TERM,
                sc.GOOD_DESCRIPTION AS SHIPMENT_TERM,
                sc.TOTAL_WEIGHT,

                cust.NAME AS BUYER_NAME,
                cust.ADDRESS AS BUYER_ADDRESS,
                cust.REPRESENTED AS LEGAL_REPRESENTATIVE,
                cust.PHONE,

                ss.EXPORT_PORT,
                ss.IMPORT_PORT,
                ss.SHIP_NAME,
                ss.ETD_DATE,
                ss.AVAILABLE_CONTAINER_QUANTITY,

                sg.UNIT_PRICE,
                g.DESCRIPTION AS COMMODITY,
                g.HS_CODE,
                g.ORIGIN_COUNTRY

            FROM T_PURCHASE_CONTRACT pc
            LEFT JOIN T_SALE_CONTRACT sc ON sc.CONTRACT_ID_INT = pc.ID
            LEFT JOIN M_CUSTOMER cust ON sc.CUSTOMER_ID_INT = cust.ID
            LEFT JOIN T_SALE_CONTRACT_GOOD sg ON sg.SALE_CONTRACT_ID_INT = sc.ID
            LEFT JOIN M_GOOD g ON sg.GOOD_ID_INT = g.ID
            LEFT JOIN T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE ss ON ss.PURCHASE_CONTRACT_ID_INT = pc.ID
            WHERE pc.CONTRACT_NUMBER = :contract_number
            LIMIT 1
        """)
        result = session.execute(query, {"contract_number": contract_number}).mappings().fetchone()
        if not result:
            return None

        net_weight = float(result["TOTAL_WEIGHT"] or 0)
        gross_weight = round(net_weight + 0.4, 2)
        unit_price = float(result["UNIT_PRICE"] or 0)
        amount = round(unit_price * net_weight, 2)
        containers = f"{int(result['AVAILABLE_CONTAINER_QUANTITY'])}x20'DC"

        return {
            "replacements": {
                "1": result["CONTRACT_TITLE"],
                "2": result["CONTRACT_NUMBER"],
                "3": format_date(result["CONTRACT_DATE"]),
                "4": result["BUYER_NAME"],
                "5": result["BUYER_ADDRESS"],
                "6": result["LEGAL_REPRESENTATIVE"],
                "7": result["PHONE"],
                "8": result["CONTRACT_NUMBER"],
                "9": result["PAYMENT_TERM"],
                "10": result["SHIPMENT_TERM"],
                "11": result["EXPORT_PORT"],
                "12": result["IMPORT_PORT"],
                "13": result["SHIP_NAME"],
                "14": format_date(result["ETD_DATE"]),
                "15": format_amount_text(amount),
                "16": result["COMMODITY"],
                "17": result["ORIGIN_COUNTRY"],
                "18": result["HS_CODE"],
                "19": result["SHIPMENT_TERM"],
                "20": unit_price,
                "21": net_weight,
                "22": gross_weight,
                "23": containers,
                "24": result["LC_CONTRACT_NUMBER"],
                "25": format_date(result["LC_DATE"]),
                "26": format_amount_text(amount)
            }
        }
    finally:
        session.close()


def fill_invoice_template(contract_number: str, template_path: str, output_path: str):
    # Gọi lấy dữ liệu
    data = get_invoice_data(contract_number)
    if not data:
        raise ValueError(f"Không tìm thấy hợp đồng: {contract_number}")
    replacements = data["replacements"]

    # Load file Excel
    workbook = load_workbook(template_path)
    sheet = workbook.active

    # Regex tìm placeholder dạng (1), (2), ...
    placeholder_pattern = re.compile(r'\((\d+)\)')  # chỉ lấy số bên trong

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                matches = placeholder_pattern.findall(cell.value)
                for match in matches:
                    key = match.strip()
                    if key in replacements:
                        cell.value = str(replacements[key])

    # Lưu kết quả
    workbook.save(output_path)
    return output_path
