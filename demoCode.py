from sqlalchemy import create_engine, text
from datetime import datetime
from num2words import num2words
import json

# db_url = "mysql+pymysql://training:123@56789@localhost:4306/smartwood"
db_url = "mysql+pymysql://training:123%4056789@localhost:4306/smartwood"

engine = create_engine(db_url)
with engine.connect() as conn:
    result = conn.execute("SELECT 1")
    print(result.scalar())
def format_date(dt, fmt='%d/%m/%Y'):
    return dt.strftime(fmt) if dt else None

def format_amount_text(amount: float) -> str:
    if amount is None:
        return ""
    text_en = num2words(amount, to='currency', lang='en')
    return f"US DOLLARS {text_en.upper()} ONLY"

def get_invoice_data(contract_number: str):
    with engine.connect() as conn:
        sql = text("""
        SELECT 
            pc.CONTRACT_NUMBER,
            pc.CONTRACT_DATE,
            CONCAT('INPUT ', pc.CONTRACT_NUMBER) AS CONTRACT_TITLE,

            sc.LC_CONTRACT_NUMBER,
            sc.LC_DATE,
            sc.PAYMENT_TERM,
            sc.GOOD_DESCRIPTION AS SHIPMENT_TERM,
            sc.TOTAL_WEIGHT,

            cust.NAME AS BUYER_NAME,
            cust.ADDRESS AS BUYER_ADDRESS,
            cust.REPRESENTED AS LEGAL_REPRESENTATIVE,
            cust.PHONE,

            ss.EXPORT_PORT,
            ss.IMPORT_PORT,
            ss.SHIP_NAME,
            ss.ETD_DATE,
            ss.AVAILABLE_CONTAINER_QUANTITY,

            sg.UNIT_PRICE,
            g.DESCRIPTION AS COMMODITY,
            g.HS_CODE,
            g.ORIGIN_COUNTRY

        FROM T_PURCHASE_CONTRACT pc
        LEFT JOIN T_SALE_CONTRACT sc ON sc.CONTRACT_ID_INT = pc.ID
        LEFT JOIN M_CUSTOMER cust ON sc.CUSTOMER_ID_INT = cust.ID
        LEFT JOIN T_SALE_CONTRACT_GOOD sg ON sg.SALE_CONTRACT_ID_INT = sc.ID
        LEFT JOIN M_GOOD g ON sg.GOOD_ID_INT = g.ID
        LEFT JOIN T_PURCHASE_CONTRACT_SHIPPING_SCHEDULE ss ON ss.PURCHASE_CONTRACT_ID_INT = pc.ID
        WHERE pc.CONTRACT_NUMBER = :contract_number
        LIMIT 1
        """)
        
        result = conn.execute(sql, {"contract_number": contract_number}).mappings().first()
        if not result:
            return None

        net_weight = float(result["TOTAL_WEIGHT"] or 0)
        gross_weight = round(net_weight + 0.4, 2)  # nếu có cột GROSS_WEIGHT thì thay vào đây
        unit_price = float(result["UNIT_PRICE"] or 0)
        amount = round(unit_price * net_weight, 2)
        containers = f"{int(result['AVAILABLE_CONTAINER_QUANTITY'])}x20'DC"

        return {
            "contract_title": result["CONTRACT_TITLE"],
            "contract_number": result["CONTRACT_NUMBER"],
            "contract_date": format_date(result["CONTRACT_DATE"]),
            "buyer_name": result["BUYER_NAME"],
            "buyer_address": result["BUYER_ADDRESS"],
            "legal_representative": result["LEGAL_REPRESENTATIVE"],
            "phone": result["PHONE"],
            "payment_term": result["PAYMENT_TERM"],
            "shipment_term": result["SHIPMENT_TERM"],
            "export_port": result["EXPORT_PORT"],
            "import_port": result["IMPORT_PORT"],
            "vessel": result["SHIP_NAME"],
            "etd_date": format_date(result["ETD_DATE"]),
            "commodity": result["COMMODITY"],
            "origin": result["ORIGIN_COUNTRY"],
            "hs_code": result["HS_CODE"],
            "unit_price": unit_price,
            "net_weight_mt": net_weight,
            "gross_weight_mt": gross_weight,
            "containers": containers,
            "lc_number": result["LC_CONTRACT_NUMBER"],
            "lc_date": format_date(result["LC_DATE"]),
            "note_amount_text": format_amount_text(amount)
        }

from openpyxl import load_workbook
import re

# 1. Load file Excel template
template_path = "Copy of Custom_Doc-DOCUMENTS_5855_TEMPLATE.xlsx"  # or full path
output_path = "Filled_Invoice_Template.xlsx"
workbook = load_workbook(template_path)
sheet = workbook.active

# 2. Tìm tất cả placeholder dạng (1), (2), ...
placeholder_pattern = re.compile(r'\(\d+\)')
placeholders = {}
for row in sheet.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            matches = placeholder_pattern.findall(cell.value)
            for match in matches:
                placeholders[match] = cell.coordinate

# 3. Dữ liệu từ get_invoice_data (ví dụ giả định từ DB)
print(get_invoice_data("2505/SW-MKS"))  # Replace with actual contract number
invoice_data = {
    "contract_title": "INPUT 30 CONT 09 OCT ECO 0899",
    "contract_number": "2505/SW-MKS",
    "contract_date": "10/02/2025",
    "buyer_name": "MK SOLAR CO., LTD",
    "buyer_address": "1201F USAN BLDG, 542 DOSAN-DAERO, GANGNAM-GU, SEOUL, SOUTH KOREA",
    "legal_representative": "MS J.S.LEE - CEO",
    "phone": "+82 2 553 5112",
    "payment_term": "AT 120 DAYS AFTER SIGHT",
    "shipment_term": "FAS HO CHI MINH PORT, VIETNAM",
    "export_port": "HO CHI MINH PORT, VIETNAM",
    "import_port": "POHANG PORT, SOUTH KOREA",
    "vessel": "KMTC KEELUNG 2417N",
    "etd_date": "14/02/2025",
    "commodity": "WOOD PELLET",
    "origin": "VIETNAM",
    "hs_code": "44013100",
    "unit_price": 133.5,
    "net_weight_mt": 391.14,
    "gross_weight_mt": 391.54,
    "containers": "20x20'DC",
    "lc_number": "M04NH2501NU00040",
    "lc_date": "22/01/2025",
    "note_amount_text": "US DOLLARS FIFTY TWO THOUSAND TWO HUNDRED AND SEVENTEEN, NINETEEN CENTS ONLY"
}

# 4. Map placeholder (1)... -> field trong JSON
placeholder_to_key = {
    "(1)": "contract_title",
    "(2)": "contract_number",
    "(3)": "contract_date",
    "(4)": "buyer_name",
    "(5)": "buyer_address",
    "(6)": "legal_representative",
    "(7)": "phone",
    "(8)": "contract_number",
    "(9)": "payment_term",
    "(10)": "shipment_term",
    "(11)": "export_port",
    "(12)": "import_port",
    "(13)": "vessel",
    "(14)": "etd_date",
    "(15)": "note_amount_text",
    "(16)": "commodity",
    "(17)": "origin",
    "(18)": "hs_code",
    "(19)": "shipment_term",
    "(20)": "unit_price",
    "(21)": "net_weight_mt",
    "(22)": "gross_weight_mt",
    "(23)": "containers",
    "(24)": "lc_number",
    "(25)": "lc_date",
    "(26)": "note_amount_text"
}

# 5. Gán dữ liệu vào file Excel
for placeholder, cell in placeholders.items():
    key = placeholder_to_key.get(placeholder)
    if key:
        value = invoice_data.get(key, "")
        sheet[cell] = value

# 6. Lưu file kết quả
workbook.save(output_path)
print(f"✅ File saved: {output_path}")


# ✅ Demo
if __name__ == "__main__":
    contract_number = "2505/SW-MKS"
    invoice_data = get_invoice_data(contract_number)
    print(json.dumps(invoice_data, indent=2, ensure_ascii=False))
